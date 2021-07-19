from openpyxl import *
import os
link =  "curl -J -O https://www.fantacalcio.it/Servizi/Excel.ashx?type=1'&'g=1'&'t=1621839592000'&'s=2020-21"
os.system(link)
wb = load_workbook(filename = 'Voti_Fantacalcio_Stagione_2020-21_Giornata_1.xlsx')
print(wb.sheetnames)
print(wb['Fantacalcio'])
ws = wb['Fantacalcio']
ws['C7']
for x in range(7,ws.max_row):
...        for y in range(1,101):
...            ws.cell(row=x, column=y)